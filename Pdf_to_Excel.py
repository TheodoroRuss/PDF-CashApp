from datetime import datetime
import os
import re
import pdfplumber
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import webbrowser


APP_TITLE = "PDF invoices - CashApp"
APP_VERSION = "v.1.0"
help_url = "https://ayahealthcare.sharepoint.com/:f:/s/Team-CostaRica-IT/IgBhCwv2QkFlQ7bjSL5vt63RAVi_DlbBWWBlL2ClK3AcJ_8?e=qYqTDA"


# ==========================================================
#  DATA EXTRACTION FUNCTIONS
# ==========================================================

def extract_invoice_data_from_text(text: str):
    """
    Extracts ONLY remittance header lines that contain:
      - Invoice #
      - Paid Invoice Amount (USD)  (supports negative values)
      - Invoice Date (YYYY-MM-DD)

    IMPORTANT:
    - Line-based matching avoids false positives caused by DOTALL across blocks.
    - Invoice Number is kept as TEXT (supports letters, hyphens, en/em dashes).
    """
    line_re = re.compile(
        r"Invoice\s*#\s*:?\s*([A-Za-z0-9][A-Za-z0-9\-–—]*)\s+"
        r"Paid\s+Invoice\s+Amount\s*:?\s*(-?[\d,]+\.\d{1,2})\s*USD\s+"
        r"Invoice\s+Date\s*:?\s*([0-9]{4}-[0-9]{2}-[0-9]{2})",
        flags=re.IGNORECASE
    )

    results = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        m = line_re.search(line)
        if not m:
            continue

        inv_num, amount, inv_date = m.groups()
        inv_num = inv_num.strip().replace("–", "-").replace("—", "-")
        amount = amount.replace(",", "").strip()
        inv_date = inv_date.strip()

        results.append((inv_num, amount, inv_date))

    return results


def extract_header_info(text: str):
    header = {"payer": None, "payment_number": None, "payment_amount": None, "credit_amount": None}
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    # Payer (name on the line after the row that starts with "Payer")
    for i, line in enumerate(lines):
        if re.search(r"\bCredit\s+Amount\b", line, flags=re.IGNORECASE):
            if i + 1 < len(lines):
                m_credit = re.search(r"([\d,]+\.\d{2})\s*USD", lines[i + 1], flags=re.IGNORECASE)
                if m_credit:
                    header["credit_amount"] = m_credit.group(1).replace(",", "").strip()
            break

    # Payment Amount
    m_amount = re.search(r"Payment\s*Amount:\s*([\d,]+\.\d{2})", text, flags=re.IGNORECASE)
    if m_amount:
        header["payment_amount"] = m_amount.group(1).replace(",", "").strip()

    # Payment Number
    m_number = re.search(r"Payment\s*Number:\s*([A-Za-z0-9\-]+)", text, flags=re.IGNORECASE)
    if m_number:
        header["payment_number"] = m_number.group(1).strip()

    return header


def extract_from_pdf(pdf_path: str, debug_print: bool = True):
    full_text = ""
    header_text = ""

    with pdfplumber.open(pdf_path) as pdf:
        if pdf.pages:
            header_text = pdf.pages[0].extract_text() or ""

        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if text:
                if debug_print:
                    print(f"\n--- Page {i} ---")
                    print(text)
                full_text += text + "\n"

    invoice_data = extract_invoice_data_from_text(full_text)
    header_info = extract_header_info(header_text)
    return invoice_data, header_info


# ==========================================================
#  EXCEL HELPER: AUTO-FIT COLUMNS
# ==========================================================

def autofit_columns(ws):
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


# ==========================================================
#  TKINTER GRAPHICAL INTERFACE (SMALL/NORMAL WINDOW)
# ==========================================================

class InvoiceExtractorApp:
    
    def reset_ui(self):
        # Reset selected file state
        self.pdf_path = None
        self.lbl_file.config(text="No file selected", fg="#8a8a8a")
        self.btn_process.config(state=tk.DISABLED)

    def __init__(self, master):
        self.master = master
        self.master.title(APP_TITLE)
        self.master.geometry("520x300")     # Normal size
        self.master.configure(bg="white")
        self.master.resizable(False, False)

        self.pdf_path = None

        # Fonts (smaller)
        self.font_title = ("Segoe UI", 18, "bold")
        self.font_sub = ("Segoe UI", 10)
        self.font_btn = ("Segoe UI", 12, "bold")
        self.font_ver = ("Segoe UI", 9)

        # Container
        container = tk.Frame(master, bg="white")
        container.pack(expand=True, fill="both", padx=20, pady=15)

        # Title
        tk.Label(
            container,
            text=APP_TITLE,
            font=self.font_title,
            bg="white",
            fg="#6b4a2b"
        ).pack(pady=(5, 8))

        # File label
        self.lbl_file = tk.Label(
            container,
            text="No file selected",
            font=self.font_sub,
            bg="white",
            fg="#8a8a8a",
            wraplength=480
        )
        self.lbl_file.pack(pady=(0, 18))

        # Buttons (smaller, same style)
        self.btn_select = tk.Button(
            container,
            text="Select PDF",
            font=self.font_btn,
            width=22,
            height=2,
            bg="white",
            fg="#333333",
            activebackground="white",
            activeforeground="#333333",
            relief="solid",
            bd=2,
            highlightthickness=0,
            command=self.select_pdf
        )
        self.btn_select.pack(pady=(0, 12))

        self.btn_process = tk.Button(
            container,
            text="Process PDF",
            font=self.font_btn,
            width=22,
            height=2,
            bg="white",
            fg="#333333",
            activebackground="white",
            activeforeground="#333333",
            relief="solid",
            bd=2,
            highlightthickness=0,
            state=tk.DISABLED,
            command=self.process_pdf
        )
        self.btn_process.pack(pady=(0, 5))

        # Version bottom-right
        tk.Label(
            master,
            text=APP_VERSION,
            font=self.font_ver,
            bg="white",
            fg="#8a8a8a"
        ).place(relx=1.0, rely=1.0, x=-10, y=-8, anchor="se")

          # Help & Latest Version (clickable link) - bottom-left
        
        self.lbl_help = tk.Label(
            master,
            text="Help & Latest Version",
            font=self.font_ver,
            bg="white",
            fg="#1a73e8",      # link-like blue
            cursor="hand2"
        )
        self.lbl_help.place(relx=0.0, rely=1.0, x=10, y=-8, anchor="sw")
        self.lbl_help.bind("<Button-1>", lambda e: webbrowser.open(help_url))

    def select_pdf(self):
        file_path = filedialog.askopenfilename(
            title="Select PDF File",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not file_path:
            return

        self.pdf_path = file_path
        self.lbl_file.config(text=os.path.basename(file_path), fg="#6b6b6b")
        self.btn_process.config(state=tk.NORMAL)

    def process_pdf(self):
        if not self.pdf_path:
            messagebox.showwarning("Warning", "Please select a PDF file first.")
            return

        try:
            wb = None
            ws = None
            data, header = extract_from_pdf(self.pdf_path, debug_print=True)

            if not data:
                messagebox.showinfo("No Data", "No invoice data found in the PDF.")
                return

            df = pd.DataFrame(data, columns=["Invoice Number", "Paid Invoice Amount", "Invoice Date"])
            df["Invoice Number"] = df["Invoice Number"].astype(str)
           # Convert amounts safely (supports negative values now)
            df["Paid Invoice Amount"] = pd.to_numeric(df["Paid Invoice Amount"], errors="coerce")

            # Drop exact duplicate invoice numbers to avoid double counting
            df = df.drop_duplicates(subset=["Invoice Number"], keep="first")

            total_amount = df["Paid Invoice Amount"].sum()

            payment_amount_header = header.get("payment_amount")
            credit_amount_header = header.get("credit_amount")

            # Try to enforce that Excel TOTAL matches Payment Amount (if present).
            # If invoices sum matches Credit Amount but not Payment Amount, it means the PDF has a net difference
            # not represented as an invoice row (e.g., fee/adjustment). We add a synthetic adjustment row.
            if payment_amount_header:
                try:
                    payment_amt = float(payment_amount_header)
                    diff = payment_amt - float(total_amount)

                    if abs(diff) >= 0.01:
                        # If invoices match Credit Amount, mismatch is structural (Credit vs Payment).
                        # Add a traceable adjustment row so the Excel TOTAL equals Payment Amount.
                        if credit_amount_header and abs(float(credit_amount_header) - float(total_amount)) < 0.01:
                            df = pd.concat(
                                [
                                    df,
                                    pd.DataFrame([{
                                        "Invoice Number": "PAYMENT_ADJUSTMENT",
                                        "Paid Invoice Amount": diff,
                                        "Invoice Date": ""
                                    }])
                                ],
                                ignore_index=True
                            )
                            total_amount = df["Paid Invoice Amount"].sum()
                except ValueError:
                    pass

            # Suggested filename = PDF name + timestamp
            pdf_filename = os.path.basename(self.pdf_path)
            pdf_name_no_ext = os.path.splitext(pdf_filename)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            suggested_name = f"{pdf_name_no_ext}_{timestamp}.xlsx"

            save_path = filedialog.asksaveasfilename(
                title="Save Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=suggested_name
            )
            if not save_path:
                return

            df.to_excel(save_path, index=False)

            wb = load_workbook(save_path)
            ws = wb.active

            # Rename sheet using payment number (fallback to 'Invoices' if missing)
            payment_number = header.get("payment_number")
            sheet_name = payment_number if payment_number else "Invoices"
            ws.title = sheet_name[:31]

            data_last_row = ws.max_row

            # Format columns
            for row in range(2, data_last_row + 1):
                inv_cell = ws.cell(row=row, column=1)
                if inv_cell.value is not None:
                    inv_cell.value = str(inv_cell.value).strip()
                inv_cell.number_format = "@"

                amt_cell = ws.cell(row=row, column=2)
                try:
                    amt_val = "" if amt_cell.value is None else str(amt_cell.value).replace(",", "").strip()
                    if amt_val != "":
                        amt_cell.value = float(amt_val)
                except Exception:
                    pass
                amt_cell.number_format = "#,##0.00"

            # TOTAL row
            last_row = data_last_row + 1
            ws.cell(row=last_row, column=1, value="TOTAL")
            total_cell = ws.cell(row=last_row, column=2, value=float(total_amount))
            total_cell.number_format = "#,##0.00"

            # Style total
            bold_font = Font(bold=True, color="FFFFFF")
            green_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col in range(1, 3):
                cell = ws.cell(row=last_row, column=col)
                cell.font = bold_font
                cell.fill = green_fill
                cell.border = border

            autofit_columns(ws)
            wb.save(save_path)
            wb.close()

            # Compare totals
            payment_amount_header = header.get("payment_amount")
            credit_amount_header = header.get("credit_amount")

            comparison_msg = ""
            if payment_amount_header:
                try:
                    header_amount = float(payment_amount_header)  # Payment Amount
                    credit_header_amount = (
                        float(credit_amount_header) if credit_amount_header else None
                    )  # Credit Amount

                    # Keep original behavior/messages, just fix the Credit Amount value shown
                    if credit_header_amount is not None:
                        if abs(header_amount - total_amount) < 0.01 and abs(credit_header_amount - total_amount) < 0.01:
                            comparison_msg = (
                                f"\n\n✅ Excel TOTAL ({total_amount:,.2f}) matches Payment Amount ({header_amount:,.2f}) and Credit Amount ({credit_header_amount:,.2f})."
                            )
                        else:
                            comparison_msg = (
                                f"\n\n⚠️ WARNING:\n"
                                f"  - Excel TOTAL: {total_amount:,.2f}\n"
                                f"  - Payment Amount (PDF): {header_amount:,.2f}\n"
                                f"  - Credit Amount (PDF): {credit_header_amount:,.2f}\n"
                                f"Values do not match."
                            )
                    else:
                        # If Credit Amount is missing, keep the same message structure as much as possible
                        if abs(header_amount - total_amount) < 0.01:
                            comparison_msg = (
                                f"\n\n✅ Excel TOTAL ({total_amount:,.2f}) matches Payment Amount ({header_amount:,.2f}) and Credit Amount ({credit_header_amount:,.2f})."
                            )
                        else:
                            comparison_msg = (
                                f"\n\n⚠️ WARNING:\n"
                                f"  - Excel TOTAL: {total_amount:,.2f}\n"
                                f"  - Payment Amount (PDF): {header_amount:,.2f}\n"
                                f"  - Credit Amount (PDF): {credit_header_amount:,.2f}\n"
                                f"Values do not match."
                            )

                except ValueError:
                    comparison_msg = "\n\nCould not interpret Payment Amount as a number."
            else:
                comparison_msg = "\n\nNo Payment Amount found in PDF."

            messagebox.showinfo(
                "Success",
                f"Data extracted successfully.\nFile saved at:\n{save_path}" + comparison_msg
            )

             # Release references (helps GC) + reset UI
            df = None
            wb = None
            ws = None
            data = None
            header = None
            self.reset_ui()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during processing:\n{e}")
        finally:
            # Ensure workbook is closed to avoid locking the output file on Windows
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass


# ==========================================================
#  MAIN
# ==========================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceExtractorApp(root)
    root.mainloop()
